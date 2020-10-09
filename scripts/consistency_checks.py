# -*- coding: utf-8 -*-
__title__ = 'histTags2mpt'
__description__ = 'to evaluate a HDSR FEWS-config with a csv with CAW histTags'
__version__ = '0.1'
__author__ = 'Daniel Tollenaar'
__author_email__ = 'daniel@d2hydro.nl'
__license__ = 'MIT License'

'''
ToDo:
    - instellingen verplaatsen naar config.ini
    - logging ook in bestand opslaan
'''

import configparser
from fews_utilities import Config, xml_to_dict
from pathlib import Path
import numpy as np
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import sys
import shutil
import re

pd.options.mode.chained_assignment = None

#%% instellingen
# layout excel spreadsheet
fixed_sheets = ['histTag_ignore',
                'inhoudsopgave',
                'exLoc_ignore',
                'TS800_ignore']

warning_sheets = ['histTags_noMatch',
                  'histTags_ignore_match',
                  'dubbele idmaps',
                  'idmap v sectie',
                  'exPar error',
                  'exPar missing',
                  'intLoc missing',
                  'exLoc error',
                  'timeSeries error']

idmap_files = ['IdOPVLWATER',
              'IdOPVLWATER_HYMOS',
              'IdHDSR_NSC',
              'IdOPVLWATER_WQ',
              'IdGrondwaterCAW']

# secties in idmap files
idmap_sections = {'IdOPVLWATER':{'KUNSTWERKEN':[{'section_start': '<!--KUNSTWERK SUBLOCS (old CAW id)-->',
                                                 'section_end': '<!--WATERSTANDSLOCATIES (old CAW id)-->'},
                                                {'section_start': '<!--KUNSTWERK SUBLOCS (new CAW id)-->',
                                                 'section_end':'<!--WATERSTANDSLOCATIES (new CAW id)-->'}],
                          'WATERSTANDLOCATIES':[{'section_start': '<!--WATERSTANDSLOCATIES (old CAW id)-->',
                                                 'section_end': '<!--MSW (old CAW id)-->'},
                                                {'section_start': '<!--WATERSTANDSLOCATIES (new CAW id)-->',
                                                'section_end': '<!--MSW (new CAW id)-->'}],
                          'MSWLOCATIES':[{'section_start': '<!--MSW (new CAW id)-->'}]}
                  }

# exParameters per sub-loc type

expars_allowed = {'pompvijzel': ['FQ.$', 'I.B$', 'IB.$', 'I.H$', 'IH.$', 'I.L$', 'IL.$', 'Q.$'],
                  'stuw': ['SW.$', 'Q.$'],
                  'schuif': ['ES.$', 'SP.$', 'SS.$', 'Q.$'],
                  'vispassage': ['ES.$', 'SP.$', 'SS.$', 'Q.$'],
                  'krooshek': ['HB.$', 'HO.$'],
                  'waterstand': ['HB.$', 'HO.$', 'H$']}

#%% functies
def idmap2tags(row,idmap):
    '''functie voor het toevoegen van fews-locatie-ids aan de hist_tags data-frame in de apply-method'''
     
    exloc, expar = row['serie'].split('_',1)
    fews_locs = [col['internalLocation'] 
                   for col in idmap
                   if col['externalLocation'] == exloc 
                   and col['externalParameter'] == expar]
    
    if len(fews_locs) == 0:
        fews_locs = np.NaN   

    return fews_locs

def update_hlocs(row):
    '''functie voor het toevoegen van start en end-date op data-frame van hoofdlocaties in de apply-method'''
    
    loc_id = row.name
    start_date = row['STARTDATE']
    end_date = row['ENDDATE']
    
    if loc_id in h_locs:
        start_date = mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]['STARTDATE'].dropna().min()
        end_date = mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]['ENDDATE'].dropna().max()
    
    return start_date, end_date 

#%% initialisatie
workdir = Path(__file__).parent
logging.basicConfig(level=os.environ.get("LOGLEVEL", "INFO"))
summary = dict()

#inlezen paden vanuit inifile
ini_config = configparser.ConfigParser()
ini_config.read(workdir.joinpath(r'..\config\config.ini'))
consistency_in = Path(r'{}'.format(ini_config['paden']['consistency_in']))
consistency_out = Path(r'{}'.format(ini_config['paden']['consistency_out']))
hist_tags_csv = Path(r'{}'.format(ini_config['paden']['hist_tags_csv']))
fews_config = Path(r'{}'.format(ini_config['paden']['fews_config']))
csv_out = Path(r'{}'.format(ini_config['paden']['csv_out']))

paths = [consistency_in, consistency_out.parent, hist_tags_csv, fews_config, csv_out]

if 'mpt_ignore' in ini_config['paden'].keys():
    mpt_ignore = Path(r'{}'.format(ini_config['paden']['mpt_ignore']))
    paths += [mpt_ignore]
else:
    mpt_ignore = None

#controleren of paden bestaan
for idx, path in enumerate(paths):
    if not path.is_absolute():
        path = workdir.joinpath(path).resolve()
        paths[idx] = path
    if not path.exists():
        if path.suffix == '':
            logging.warning(f'{path} bestaat niet, map wordt aangemaakt')
            path.mkdir()
        else:
            logging.error(f'{path} bestaat niet. Specificeer het juiste path in config.ini')
            sys.exit()

#%% inlezen config-excel
# kopieeren van consistency workbook naar output
try:
    shutil.copyfile(consistency_in, consistency_out)
except Exception as e: 
    logging.error(e) 
    sys.exit()

config_df = pd.read_excel(consistency_in,sheet_name=None,engine='openpyxl')
if not 'histTag_ignore' in config_df.keys():
    logging.error('werkblad "histTag_ignore" mist in {}'.format(consistency_in))
    sys.exit()
    
# weggooien van alle output-sheets, behalve degenen opgeschoond
config_df = {key:value for key,value in config_df.items() if key in fixed_sheets}

#%% inlezen idmap-files
config = Config(fews_config)
idmap_dict = {idmap:xml_to_dict(config.IdMapFiles[idmap])['idMap']['map'] 
           for idmap in idmap_files}
idmap_total = [j for i in idmap_dict.values() for j in i]
    
#%% controle op KW/OW
logging.info('controle op KW/OW locaties in juiste sectie')
config_df['idmap v sectie'] = pd.DataFrame(columns=['bestand',
                                                    'externalLocation',
                                                    'externalParameter',
                                                    'internalLocation',
                                                    'internalParameter',
                                                    ])
for idmap, idmap_subsecs in idmap_sections.items():
    for section_type, sections in idmap_subsecs.items():
        for section in sections:
            if section_type == 'KUNSTWERKEN':
                prefix = 'KW'
            if section_type == 'WATERSTANDLOCATIES':
                prefix = 'OW'
            if section_type == 'MSWLOCATIES':
                prefix = '(OW|KW)'
            pattern = f'{prefix}\d{{6}}$'
            idmap_wrong_section = [idmap for idmap in xml_to_dict(config.IdMapFiles[idmap],**section)['idMap']['map'] 
                                   if not bool(re.match(pattern,idmap['internalLocation']))]
            if len(idmap_wrong_section):
                section_start = section['section_start'] if 'section_start' in section.keys() else ''
                section_end = section['section_end'] if 'section_end' in section.keys() else ''
                logging.warning('{} internalLocations anders dan {}XXXXXX tussen {} en {} in {}'.format(len(idmap_wrong_section),
                                                                                                       prefix,
                                                                                                       section_start,
                                                                                                       section_end,
                                                                                                       idmap))
                df = pd.DataFrame(idmap_wrong_section)
                df['sectie'] = section_start
                df['bestand'] = idmap
                config_df['idmap v sectie'] = pd.concat([config_df['idmap v sectie'], df], axis=0)
    
    summary['idmaps in verkeerde sectie'] = len(config_df['idmap v sectie'])

#%% inlezen hist tags & ignore lijst
logging.info('zoeken naar missende histTags in idmaps')
dtype_cols = ['total_min_start_dt', 'total_max_end_dt']
hist_tags_org_df = pd.read_csv(hist_tags_csv,
                           parse_dates = dtype_cols,
                           sep = ';')

for col in dtype_cols:
    if not pd.api.types.is_datetime64_dtype(hist_tags_org_df[col]):
        logging.error(f"kolom '{col}' in '{hist_tags_csv}' kan niet worden geconverteerd"
                      " naar np.datetime64 formaat. Controleer of deze datums realistisch zijn.")
        sys.exit()

#%% filteren hist_tags op alles wat niet in ignored staat
hist_tags_df = hist_tags_org_df.copy()
hist_tags_df['fews_locid'] = hist_tags_org_df.apply(idmap2tags, args=[idmap_total], axis=1)
hist_tags_no_match_df = hist_tags_df[hist_tags_df['fews_locid'].isna()]
hist_tags_no_match_df = hist_tags_no_match_df[~hist_tags_no_match_df['serie'].isin(config_df['histTag_ignore']['UNKNOWN_SERIE'])] 
hist_tags_no_match_df = hist_tags_no_match_df.drop('fews_locid',axis=1)
hist_tags_no_match_df.columns = ['UNKNOWN_SERIE','STARTDATE','ENDDATE']
hist_tags_no_match_df = hist_tags_no_match_df.set_index('UNKNOWN_SERIE')
config_df['histTags_noMatch'] = hist_tags_no_match_df
summary['histTags_noMatch'] = len(hist_tags_no_match_df)

if not config_df['histTags_noMatch'].empty:
    logging.warning('{} histTags zijn niet opgenomen in idmap'.format(len(config_df['histTags_noMatch'])))
else:
    logging.info('alle histTags zijn opgenomen in idmap')

#%% wegschrijven van ids die ten onrechte in ignore-lijst staan
if mpt_ignore:
   config_df['histTag_ignore'] = pd.read_csv(mpt_ignore,sep=';',header=0)  
config_df['histTag_ignore']['UNKNOWN_SERIE'] = config_df['histTag_ignore']['UNKNOWN_SERIE'].str.replace('#','')   
hist_tags_opvlwater_df = hist_tags_org_df.copy()
hist_tags_opvlwater_df['fews_locid'] = hist_tags_org_df.apply(idmap2tags, args=[idmap_dict['IdOPVLWATER']], axis=1)
hist_tags_opvlwater_df = hist_tags_opvlwater_df[hist_tags_opvlwater_df['fews_locid'].notna()]
hist_tag_ignore_match_df = config_df['histTag_ignore'][config_df['histTag_ignore']['UNKNOWN_SERIE'].isin(hist_tags_opvlwater_df['serie'])]
hist_tag_ignore_match_df = hist_tag_ignore_match_df.set_index('UNKNOWN_SERIE')
config_df['histTags_ignore_match'] = hist_tag_ignore_match_df

if not config_df['histTags_ignore_match'].empty:
    logging.warning('{} histTags zijn ten onrechte opgenomen in histTag ignore'.format(len(config_df['histTags_ignore_match'])))
else:
    logging.info('geen histTags ten onrechte in ignore')

#%% aanmaken van mpt_df vanuit de fews_locid lijsten in hist_tags_df
logging.info('omzetten van histTags naar meetpunten')
hist_tags_df = hist_tags_df[hist_tags_df['fews_locid'].notna()]
mpt_hist_tags_df = hist_tags_df.explode('fews_locid').reset_index(drop=True)

# bepalen minimale start en maximale eindtijd per fews_locid. 
mpt_df = pd.concat([mpt_hist_tags_df.groupby(['fews_locid'], sort=False)['total_min_start_dt'].min(),
                    mpt_hist_tags_df.groupby(['fews_locid'], sort=False)['total_max_end_dt'].max()],
                   axis=1)

mpt_df = mpt_df.sort_index(axis=0)
mpt_df.columns = ['STARTDATE','ENDDATE']
mpt_df.index.name = 'LOC_ID'

# alle hoofdlocaties waar geen histag op binnekomt toevoegen
kw_locs = list(mpt_df[mpt_df.index.str.contains('KW', regex=False)].index)
h_locs = np.unique(['{}0'.format(loc[0:-1]) for loc in kw_locs])
h_locs_missing = [loc for loc in h_locs if not loc in list(mpt_df.index)]
h_locs_df = pd.DataFrame(data={'LOC_ID' : h_locs_missing,
                               'STARTDATE' : [pd.NaT]*len(h_locs_missing),
                               'ENDDATE' :  [pd.NaT]*len(h_locs_missing)})
h_locs_df = h_locs_df.set_index('LOC_ID')

mpt_df = pd.concat([mpt_df,h_locs_df],axis=0)
# de start en eindtijd op de hoofdlocatie updaten met de min/max van de sublocatie

mpt_df[['STARTDATE','ENDDATE']] = mpt_df.apply(update_hlocs,axis=1,result_type="expand")

mpt_df = mpt_df.sort_index()
config_df['mpt'] = mpt_df

#%% consistentie parameters: zijn alle interne parameters opgenomen in parameters.xml
logging.info('controle dubbele idmaps')
config_df['dubbele idmaps'] = pd.DataFrame(columns=['bestand',
                                                    'externalLocation',
                                                    'externalParameter',
                                                    'internalLocation',
                                                    'internalParameter'])                    
for idmap_file in idmap_files:
    idmap_doubles = [id_map for id_map in idmap_dict[idmap_file] if idmap_dict[idmap_file].count(id_map) > 1]
    if len(idmap_doubles) > 0:
        idmap_doubles = list({idmap['externalLocation']:idmap for idmap in idmap_doubles}.values())
        df = pd.DataFrame(idmap_doubles,columns=['internalLocation','externalLocation','internalParameter','externalParameter'])
        df['bestand'] = idmap_file
        config_df['dubbele idmaps'] = pd.concat([config_df['dubbele idmaps'], df], axis=0)
        logging.warning('{} dubbele idmap(s) in {}'.format(len(idmap_doubles),idmap_file))
    else:
        logging.info('geen dubbele idmaps in {}'.format(idmap_file))

    summary['dubbele idmaps {}'.format(idmap_file)] = len(idmap_doubles)

#%% consistentie parameters: zijn alle interne parameters opgenomen in parameters.xml
config_parameters = list(config.get_parameters(dict_keys='parameters').keys())
id_map_parameters = [id_map['internalParameter'] for id_map in idmap_total]
params_missing = [parameter for parameter in id_map_parameters 
                  if not parameter in config_parameters]

summary['missende parameters'] = len(params_missing)

if len(params_missing) == 0:
    logging.info('alle parameters in idMaps zijn opgenomen in config')
else:
    logging.warning('{} uit idMaps missen in config'.format(len(params_missing)))
    config_df['params_missing'] =  pd.DataFrame({'parameters': params_missing})
    config_df['params_missing'] = config_df['params_missing'].set_index('parameters')

#%% consistentie externe parameters met interne parameters/locaties
logging.info('controle foutieve en missende ex-parameters & niet opgenomen inlocs')
hoofdloc_gdf = config.get_locations('OPVLWATER_HOOFDLOC')
subloc_gdf = config.get_locations('OPVLWATER_SUBLOC')
waterstand_gdf = config.get_locations('OPVLWATER_WATERSTANDEN_AUTO')
msw_gdf = config.get_locations('MSW_STATIONS')

ex_par_errors = {'internalLocation':[],
                 'locationType':[],
                 'exParError':[],
                 'types':[],
                 'FQ':[],
                 'I.X':[],
                 'IX.':[]}

ex_par_missing = {'internalLocation':[],
                  'locationType':[],
                  'exPars':[],
                  'QR':[],
                  'QS':[],
                  'HS':[]}

int_loc_missing = []

#maak een data-frame zodat we kunnen groeperen bij internalLocation
idmap_df = pd.DataFrame.from_dict(idmap_dict['IdOPVLWATER'])

for loc_group in idmap_df.groupby('internalLocation'):
    #initieer een aantal variabelen
    missings = dict.fromkeys(['QR','QS','HS'],False)
    errors = dict.fromkeys(['I.X','IX.','FQ'],False)
    
    #interne locatie en externe parameters
    int_loc = loc_group[0]
    ex_pars = np.unique(loc_group[1]['externalParameter'].values)
    ex_pars_gen = [re.sub("\d", ".", ex_par) for ex_par in ex_pars]
    
    #vaststellen locatie-type
    if int_loc in hoofdloc_gdf['LOC_ID'].values:
        loc_properties = hoofdloc_gdf[hoofdloc_gdf['LOC_ID'] == int_loc]
        loc_type = 'hoofdloc'
    elif int_loc in subloc_gdf['LOC_ID'].values:
        loc_properties = subloc_gdf[subloc_gdf['LOC_ID'] == int_loc]
        loc_type = 'subloc'
        regexes = ['HR.$']
    elif int_loc in waterstand_gdf['LOC_ID'].values:
        loc_type = 'waterstand'
    elif int_loc in msw_gdf['LOC_ID'].values:
        loc_type = 'msw'
    else:
        loc_type = None
        int_loc_missing += [int_loc]
    
    #vaststellen object_typen
    if loc_type in ['hoofdloc', 'subloc']:
        all_types = loc_properties['ALLE_TYPES'].values[0].split("/")
        all_types = [item.lower() for item in all_types]
    elif loc_type == 'waterstand':
        all_types = ['waterstand']
    
    if loc_type == 'subloc':
        
        #zoeken naar foutief toegekende ex_pars
        regexes += [j for i in
                    [values for keys, values in expars_allowed.items() if keys in all_types]
                    for j in i]
        regexes += list(dict.fromkeys(regexes))
        
        ex_par_error = [ex_par for ex_par in ex_pars if not any([regex.match(ex_par) for regex in [re.compile(rex) for rex in regexes]])]
         
        # als wel/niet I.B dan ook wel/niet IB.
        if any([ex_par for ex_par in ex_pars_gen if ex_par in ['I.B', 'I.H', 'I.L']]):
            if not any([ex_par for ex_par in ex_pars_gen if ex_par in ['IB.', 'IH.', 'IL.']]):
                errors['IX.'] = True
        elif any([ex_par for ex_par in ex_pars_gen if ex_par in ['IB.', 'IH.', 'IL.']]):
             errors['I.X'] = True
        
        # Als FQ, dan ook I.B.
        if 'FQ.' in ex_pars_gen: 
            if not any([ex_par for ex_par in ex_pars_gen if ex_par in ['IB.', 'IH.', 'IL.', 'I.B', 'I.H', 'I.L']]):
                errors['FQ'] = True
                
    elif loc_type == 'hoofdloc':
        
        #zoeken naar foutief toegekende ex_pars
        regexes = ['HS.$', 'QR.$', 'QS.$', 'WR', 'WS']
        
        ex_par_error = [ex_par for ex_par in ex_pars if not any([regex.match(ex_par) for regex in [re.compile(rex) for rex in regexes]])]
        
        #is er een HS?
        if not ('HS.' in ex_pars_gen):
            missings['HS'] = True
            
        if not ('QR.' in ex_pars_gen):
            missings['QR'] = True
            
        if not ('QS.' in ex_pars_gen):
            missings['QS'] = True
            
    else:
        ex_par_error = []
            
    # rapporteren expar_errors
    if len(ex_par_error) > 0 | any(errors.values()):
        ex_par_errors['internalLocation'].append(int_loc)
        ex_par_errors['locationType'].append(loc_type)
        ex_par_errors['exParError'].append(','.join(ex_par_error))
        ex_par_errors['types'].append(','.join(all_types))
        for key, value in errors.items():
            ex_par_errors[key].append(value)
        
    # rapporteren missings
    if any(missings.values()):
        ex_par_missing['internalLocation'].append(int_loc)
        ex_par_missing['locationType'].append(loc_type)
        ex_par_missing['exPars'].append(','.join(ex_pars))
        for key, value in missings.items():
            ex_par_missing[key].append(value)
    
#opname in data-frame           
config_df['exPar error'] = pd.DataFrame(ex_par_errors)
config_df['exPar missing'] = pd.DataFrame(ex_par_missing)
config_df['intLoc missing'] = pd.DataFrame({'internalLocation':int_loc_missing})

#opname in samenvatting
summary['ExPar errors'] = len(config_df['exPar error'])
summary['ExPar missing'] = len(config_df['exPar missing'])
summary['IntLoc missing'] = len(config_df['intLoc missing'])

#loggen van resultaat
for item in ['ExPar errors', 'ExPar missing']:
    if summary[item] == 0:
        logging.info('geen {}'.format(item))
    else:
      logging.warning('{} locaties met {}'.format(summary[item],item))
      
if summary['IntLoc missing'] == 0:
    logging.info('alle interne locaties uit idmap opgenomen in locationSets')
else:
    logging.warning('{} interne locaties niet opgenomen in locationSets'.format(summary['IntLoc missing']))                    

#%% zoeken naar ex-loc errors
logging.info('controle externe locaties')
ex_loc_errors = {'internalLocation':[],
                 'externalLocation':[]}   

for loc_group in idmap_df.groupby('externalLocation'):

    #initialiseren int_loc_error
    int_loc_error = []
    
    #zoeken naar ex-loc errors
    ex_loc = loc_group[0]
    int_locs = np.unique(loc_group[1]['internalLocation'].values)
    
    # als lengte van ex-loc == 3
    if len(ex_loc) == 3:
        
        # de default-case
        if not bool(re.match('8..$',ex_loc)):
            int_loc_error = [int_loc for int_loc in int_locs if 
                             not bool(re.match(f'...{ex_loc}..$',int_loc))]
         
        # opgesplitste locaties; ex-loc altijd naar 1 unieke hoofdlocatie + sublocaties
        else:
            for loc_type in ['KW','OW']:
                int_locs_select = [int_loc for int_loc in int_locs 
                                   if bool(re.match(f'{loc_type}.',int_loc))]
                if len(np.unique([int_loc[:-1] for int_loc in int_locs_select])) > 1:
                    int_loc_error += list(int_locs_select)
    
    # als lengte ex-loc == 4
    if len(ex_loc) == 4:
        
        # de default-case
        if not bool(re.match('.8..$',ex_loc)):
            int_loc_error += [int_loc for int_loc in int_locs if 
                              not bool(re.match(f'..{ex_loc}..$',int_loc))]
        
        # opgesplitste locaties; ex-loc altijd naar 1 unieke hoofdlocatie + sublocaties
        else:
            for loc_type in ['KW','OW']:
                int_locs_select = [int_loc for int_loc in int_locs 
                                   if bool(re.match(f'{loc_type}.',int_loc))]
                if len(np.unique([int_loc[:-1] for int_loc in int_locs_select])) > 1:
                    int_loc_error += list(int_locs_select)
    
    #als de ex-loc in de ignore-lijst staan, dan int_loc_error opruimen
    if 'exLoc_ignore' in config_df.keys():
        if int(ex_loc) in config_df['exLoc_ignore']['externalLocation'].values:
            int_loc_error = [int_loc for int_loc in int_loc_error 
                               if not int_loc in 
                               config_df['exLoc_ignore'][config_df['exLoc_ignore']['externalLocation'] 
                                                         == int(ex_loc)]['internalLocation'].values]
             
    for int_loc in int_loc_error:
        ex_loc_errors['internalLocation'].append(int_loc)
        ex_loc_errors['externalLocation'].append(ex_loc)

config_df['exLoc error'] = pd.DataFrame(ex_loc_errors)

summary['exLoc error'] = len(config_df['exLoc error'])

if summary['exLoc error'] == 0:
    logging.info('alle externe locaties consistent met interne locaties')
else:
    logging.warning('{} externe locaties onlogisch bij interne locaties'.format(summary['exLoc error']))                    

#%% zoeken naar sub-locaties anders dan krooshek en debietmeter:
#   - zonder stuurpeil tijdserie
#   - waarbij meerdere tijdseries met stuurpeilen naar dezelfde interne paramer mappen
logging.info('controle koppeling tijdseries')
if 'TS800_ignore' in config_df.keys():
    ts_ignore_df = config_df['TS800_ignore']
else:
    ts_ignore_df = pd.DataFrame({'internalLocation':[],'externalLocation':[]})

#%%
idmap_subloc_df = idmap_df[idmap_df['internalLocation'].isin(subloc_gdf['LOC_ID'].values)] # alleen locaties die in de sub-locs locationSet zitten
idmap_subloc_df['type'] = idmap_subloc_df['internalLocation'].apply((lambda x: subloc_gdf[subloc_gdf['LOC_ID'] == x]['TYPE'].values[0])) #toevoegen van type

#%%
#idmap_subloc_df = idmap_subloc_df[~idmap_subloc_df['type'].isin(['krooshek','debietmeter'])] # krooshekken en debietmeters zijn niet relevant
idmap_subloc_df['loc_groep'] = idmap_subloc_df['internalLocation'].apply((lambda x: x[0:-1]))

ts_errors = {'internalLocation':[],
             'internalParameters':[],
             'externalParameters':[],
             'externalLocations':[],
             'type':[],
             'fout':[]
             }

for loc_group, group_df in idmap_subloc_df.groupby('loc_groep'):
    
    #uniek nummer per ex-loc
    ex_locs = np.unique(group_df['externalLocation'].values)
    ex_locs_dict = {ex_loc:idx for idx, ex_loc in enumerate(ex_locs)}
    
    #vinden van 800 nummers
    split_ts = [key for key in ex_locs_dict.keys() if 
                  any([regex.match(key) 
                        for regex in [re.compile(rex) 
                                      for rex in ['8..','.8..']]])]
    
    ex_locs_skip = ts_ignore_df[ts_ignore_df['internalLocation'].isin(group_df['internalLocation'])]['externalLocation']
    
    split_ts = [key for key in split_ts if not str(key) in ex_locs_skip.values.astype(np.str)]
    
    ex_locs_dict = {k:(ex_locs_dict[k[1:]] 
                        if (k[1:] in ex_locs_dict.keys()) and (not k in split_ts) 
                        else v) for (k,v) in ex_locs_dict.items()}
    
    org_uniques = np.unique([val for key,val in ex_locs_dict.items() if not key in split_ts])
       
    # als er maar 1 groep zit in split_ts én een groep in de originele tijdseriegroepen, dan samenvoegen
    if (len(org_uniques) == 1) & (len(split_ts) == 1):
        ex_locs_dict = {k:(org_uniques[0] if k in split_ts else v) for (k,v) in ex_locs_dict.items()}
        
    group_df['ex_loc_group'] = group_df['externalLocation'].apply((lambda x: ex_locs_dict[x]))
 
    for int_loc, loc_df in group_df.groupby('internalLocation'):
        loc_type = subloc_gdf[subloc_gdf['LOC_ID'] == int_loc]['TYPE'].values[0]
        ex_pars = np.unique(loc_df['externalParameter'].values)
        int_pars = np.unique(loc_df['internalParameter'].values)
        ex_locs = np.unique(loc_df['externalLocation'].values)
        
        if loc_type in ['krooshek','debietmeter']:
            if any([re.match('HR.',ex_par) for ex_par in ex_pars]):
                #krooshek/debietmeter met stuurpeil = fout
                ts_errors['internalLocation'].append(int_loc)
                ts_errors['internalParameters'].append(",".join(int_pars))
                ts_errors['externalParameters'].append(",".join(ex_pars))
                ts_errors['externalLocations'].append(','.join(ex_locs))
                ts_errors['type'].append(loc_type)
                ts_errors['fout'].append(f'{loc_type} met stuurpeil')
        
        else: #geen krooshek of debietmeter
            # geen sp, maar wel sp op andere subloc = fout
            if (not any([re.match('HR.',ex_par) for ex_par in ex_pars])): # geen stuurpeil
                if any([re.match('HR.',ex_par) for ex_par in np.unique(group_df['externalParameter'])]):
                    #~krooshek/debietmeter zonder stuurpeil = fout
                    sp_locs = np.unique(group_df[group_df['externalParameter'].str.match('HR.')]['internalLocation'])
                    ts_errors['internalLocation'].append(int_loc)
                    ts_errors['internalParameters'].append(",".join(int_pars))
                    ts_errors['externalParameters'].append(",".join(ex_pars))
                    ts_errors['externalLocations'].append(','.join(ex_locs))
                    ts_errors['type'].append(loc_type)
                    ts_errors['fout'].append(f'{loc_type} zonder stuurpeil ({",".join(sp_locs)} wel)')
                    
            else: #krooshek/debietmeter met stuurpeil
                # >1 sp zonder andere interne parameter = fout
                time_series = loc_df.groupby(['ex_loc_group','externalParameter'])
                sp_series = [series for series in time_series if bool(re.match('HR.',series[0][1]))]
                for idx, series in enumerate(sp_series):
                    ex_par = series[0][1]
                    ex_locs = series[1]['externalLocation']
                    int_par = np.unique(series[1]['internalParameter'])
                    if len(int_par) > 1:
                        # 1 sp series gekoppeld aan 2 fews parameters
                        ts_errors['internalLocation'].append(int_loc)
                        ts_errors['internalParameters'].append(",".join(int_pars))
                        ts_errors['externalParameters'].append(",".join(ex_pars))
                        ts_errors['externalLocations'].append(','.join(ex_locs))
                        ts_errors['type'].append(loc_type)
                        ts_errors['fout'].append(f'{",".join(int_par)} gekoppeld aan 1 sp-serie (exPar: {ex_par}, exLoc(s)): {",".join(ex_locs)}')
                        
                    other_series = [series for idy, series in enumerate(sp_series) if not idy == idx]
                    other_int_pars = [np.unique(series[1]['internalParameter']) for series in other_series]
                    if len(other_int_pars) > 0: other_int_pars = np.concatenate(other_int_pars)
                    conflicting_pars = [par for par in int_par if par in other_int_pars]
                    if len(conflicting_pars) > 0:
                        # 2 sp series gekoppeld aan dezelfde fews parameter
                        ts_errors['internalLocation'].append(int_loc)
                        ts_errors['internalParameters'].append(",".join(int_pars))
                        ts_errors['externalParameters'].append(",".join(ex_pars))
                        ts_errors['externalLocations'].append(','.join(ex_locs))
                        ts_errors['type'].append(loc_type)
                        ts_errors['fout'].append(f'{",".join(conflicting_pars)} gekoppeld aan sp-serie (exPar: {ex_par}, exLoc(s)): {",".join(ex_locs)}')
                        
                    
            
            
#     if len(group_df.groupby('internalLocation')) == 1:
#         int_loc = list(group_df.groupby('internalLocation'))[0][0]
#         loc_type = list(group_df.groupby('type'))[0][0]
#         sp_series = [series for series in time_series if bool(re.match('HR.',series[0][1]))]
#         unique, counts = np.unique([np.unique(series[1]['internalParameter']) for series in sp_series],return_counts=True)
#         sp_params = dict(zip(unique, counts)) 
#         for series in sp_series:
#             params = np.unique(series[1]['internalParameter'])
#             if len(params) > 1:
#                 errors['tijdseries'] = f"dezelfde reeks gekoppeld aan parameters {','.join(params)}"
#                 #logging.error(f'interne locatie {int_loc} heeft dezelfde tijdreeks gekoppeld aan interne parameters {",".join(params)}')
#             elif sp_params[params[0]] > 1:
#                 errors['tijdseries'] = f"naast exLoc {','.join(np.unique(series[1]['externalLocation']))} en exPar {series[0][1]} nog andere reeks gekoppeld"
#                 #logging.error(f"interne locatie {int_loc} heeft naast tijdreeks met exLoc {','.join(np.unique(series[1]['externalLocation']))} en exPar {series[0][1]}"
#                 #              f" nog een andere reeks gekoppeld aan {params[0]}")
#             if errors['tijdseries']:
#                 ts_errors['internalLocation'].append(int_loc)
#                 ts_errors['internalParameters'].append(",".join(int_pars))
#                 ts_errors['externalParameters'].append(",".join(ex_pars))
#                 ts_errors['externalLocations'].append(','.join(np.unique(series[1]['externalLocation'])))
#                 ts_errors['type'].append(loc_type)
#                 ts_errors['fout'].append(errors['tijdseries'])
                
#     if not any([bool(re.match('HR.',ex_par)) for ex_par in ex_pars]):
#         errors['stuurpeil'] = 'missend stuurpeil'
#         #logging.error(f'interne locatie {int_loc} van type {loc_type} heeft geen stuurpeil')
#         if errors['stuurpeil']:
#             ts_errors['internalLocation'].append(int_loc)
#             ts_errors['internalParameters'].append(",".join(int_pars))
#             ts_errors['externalParameters'].append(",".join(ex_pars))
#             ts_errors['externalLocations'].append(','.join(np.unique(series[1]['externalLocation'])))
#             ts_errors['type'].append(loc_type)
#             ts_errors['fout'].append('missend stuurpeil')
        
config_df['timeSeries error'] = pd.DataFrame(ts_errors)

#opname in samenvatting
summary['timeSeries errors'] = len(config_df['timeSeries error'])

if summary['timeSeries errors'] == 0:
    logging.info('alle tijdseries zijn logisch gekoppeld aan interne locaties/parameters')
else:
    logging.warning('{} tijdseries missend/onlogisch gekoppeld'.format(summary['exLoc error']))   

#%% controle validationrulesets
logging.info('controle validationRules')

validation_rules = xml_to_dict(config.RegionConfigFiles['ValidationRuleSets'])['validationRuleSets']['validationRuleSet']
validation_rules = [rule for rule in validation_rules if 'extremeValuesFunctions' in rule.keys()]

#for validation_rule in validation_rules:
validation_rule = validation_rules[4]
location_sets = np.unique([ts['locationSetId'] for ts in validation_rule['timeSeriesSet']])

location_set = location_sets[0]
config.locationSets[location_set]

hard_max = next(v for (k,v) in validation_rule['extremeValuesFunctions'].items() if k == 'hardMax')
soft_max = next(v for (k,v) in validation_rule['extremeValuesFunctions'].items() if k == 'softMax')
hard_min = next(v for (k,v) in validation_rule['extremeValuesFunctions'].items() if k == 'hardMin')
soft_min = next(v for (k,v) in validation_rule['extremeValuesFunctions'].items() if k == 'softMin')


#%% wegschrijven naar excel
    
#lees input xlsx en gooi alles weg behalve de fixed_sheets
book = load_workbook(consistency_out)
for worksheet in book.worksheets:
    if not worksheet.title in fixed_sheets:
        book.remove(worksheet)

# voeg samenvatting toe
worksheet = book.create_sheet('samenvatting',1)
worksheet.sheet_properties.tabColor = '92D050'
worksheet.append(['controle','aantal'])
for cell in worksheet['{}'.format(worksheet.max_row)]:
    cell.font = Font(bold=True)
    
for key, value in summary.items():
    worksheet.append([key,value])
    if value > 0:
       worksheet[worksheet.max_row][1].fill = PatternFill(fgColor='FF0000', fill_type='solid')
    else:
        worksheet[worksheet.max_row][1].fill = PatternFill(fgColor='92D050', fill_type='solid')

worksheet.column_dimensions['A'].width=40
worksheet.auto_filter.ref = worksheet.dimensions

xls_writer = pd.ExcelWriter(consistency_out, engine='openpyxl')
xls_writer.book = book

for sheet_name, df in config_df.items():
        if (not sheet_name in fixed_sheets) & (not df.empty):
            if df.index.name == None:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=True)
            worksheet = xls_writer.sheets[sheet_name]
            for col in worksheet.columns:
                worksheet.column_dimensions[col[0].column_letter].width = 20
            worksheet.auto_filter.ref = worksheet.dimensions
            if not df.empty:
                if (sheet_name in warning_sheets):
                    worksheet.sheet_properties.tabColor = 'FF0000'
                else:
                    worksheet.sheet_properties.tabColor = '92D050'
                    
xls_writer.book.active = xls_writer.book['samenvatting']

xls_writer.save()

#%% updaten csv's
def update_csv(row,mpt_df,date_threshold):
    int_loc = row['LOC_ID']
    if int_loc in mpt_df.index:
        start_date = mpt_df.loc[int_loc]['STARTDATE'].strftime('%Y%m%d')
        end_date = mpt_df.loc[int_loc]['ENDDATE']
        if end_date > date_threshold:
            end_date = pd.Timestamp(year=2100, month=1, day=1)
        end_date = end_date.strftime('%Y%m%d')
    else:
        start_date = row['START']
        end_date = row['EIND']
        
    return start_date, end_date

#%%
date_threshold = mpt_df['ENDDATE'].max() - pd.Timedelta(weeks=26)

for locationSet, gdf in {'OPVLWATER_HOOFDLOC': hoofdloc_gdf,
                         'OPVLWATER_SUBLOC': subloc_gdf,
                         'OPVLWATER_WATERSTANDEN_AUTO': waterstand_gdf}.items():
    logging.info(f'wegschrijven csv voor locationSet: {locationSet}')
    df = gdf.drop('geometry',axis=1)
    df[['START','EIND']] = df.apply(update_csv, 
                                    args=(mpt_df, date_threshold), 
                                    axis=1,
                                    result_type="expand")

    csv_file = csv_out.joinpath(config.locationSets[locationSet]['csvFile']['file'])
    if csv_file.suffix == '':
        csv_file = Path(f'{csv_file}.csv')
    df.to_csv(csv_file, index=False)