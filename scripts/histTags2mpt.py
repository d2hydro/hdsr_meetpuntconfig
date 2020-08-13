# -*- coding: utf-8 -*-
__title__ = 'histTags2mpt'
__description__ = 'to evaluate a HDSR FEWS-config with a csv with CAW histTags'
__version__ = '0.1'
__author__ = 'Daniel Tollenaar'
__author_email__ = 'daniel@d2hydro.nl'
__license__ = 'MIT License'

'''
ToDo:
'''

import configparser
from fews_utilities import Config, xml_to_dict
import numpy as np
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import sys
import shutil


# initialisatie
logging.basicConfig(level=os.environ.get("LOGLEVEL", "INFO"))
summary = dict()
config = configparser.ConfigParser()
config.read(r'..\config\config.ini')

# consistency_in = r'..\data\consistency.xlsx' #vorig resultaat-bestand
# consistency_out = r'..\data\consistency_uit.xlsx' #pad naar nieuw resultaat-bestand
# hist_tags_csv = r'..\data\get_series_startenddate_CAW_summary_total_sorted_20200405.csv' #csv met histTags
# fews_config = r'd:\FEWS\HDSR_WIS\CAW\config' #pad naar FEWS-config

# paden
consistency_in = r'{}'.format(config['paden']['consistency_in'])
consistency_out = r'{}'.format(config['paden']['consistency_out'])
hist_tags_csv = r'{}'.format(config['paden']['hist_tags_csv'])
fews_config = r'{}'.format(config['paden']['fews_config'])

# layout xlsx-file
fixed_sheets = ['histTag_ignore','inhoudsopgave']
warning_sheets = ['histTags_noMatch','dubbele idmaps','idmap v sectie']
idmap_files = ['IdOPVLWATER',
              'IdOPVLWATER_HYMOS',
              'IdHDSR_NSC',
              'IdOPVLWATER_WQ',
              'IdGrondwaterCAW']

# secties in idmap files
idmap_sections = {'IdOPVLWATER':{'KUNSTWERKEN':[{'section_start': '<!--KUNSTWERK SUBLOCS (old CAW id)-->',
                                                 'section_end': '<!--WATERSTANDSLOCATIES (old CAW id)-->'},
                                                {'section_start': '<!--KUNSTWERK SUBLOCS (new CAW id)-->',
                                                 'section_end':'<!--WATERSTANDSLOCATIES (new CAW id)-->'}],
                          'WATERSTANDLOCATIES':[{'section_start': '<!--WATERSTANDSLOCATIES (old CAW id)-->',
                                                 'section_end': '<!--KUNSTWERK SUBLOCS (new CAW id)-->'},
                                                {'section_start': '<!--WATERSTANDSLOCATIES (new CAW id)-->'}]}}



#%% functies
def idmap2tags(row):
    '''functie voor het toevoegen van fews-locatie-ids aan de hist_tags data-frame'''
     
    exloc, expar = row['serie'].split('_',1)
    fews_locs = [col['internalLocation'] 
                   for col in idmap_total 
                   if col['externalLocation'] == exloc 
                   and col['externalParameter'] == expar]
    
    if len(fews_locs) == 0:
        fews_locs = np.NaN   

    return fews_locs

#%% inlezen config-excel
try:
    shutil.copyfile(consistency_in, consistency_out)
except Exception as e: 
    logging.error(e) 
    sys.exit()

config_df = pd.read_excel(consistency_in,sheet_name=None)
if not 'histTag_ignore' in config_df.keys():
    logging.error('werkblad "histTag_ignore" mist in {}'.format(consistency_in))
    sys.exit()
    
# weggooien van alle output-sheets
config_df = {key:value for key,value in config_df.items() if key in fixed_sheets}

#%% inlezen idmap
config = Config(fews_config)
idmap_total = []
idmap_dict = {idmap:xml_to_dict(config.IdMapFiles[idmap])['idMap']['map'] 
           for idmap in idmap_files}

for idmap in idmap_dict.values():
    idmap_total += idmap
    
#%% controle op KW/OW
config_df['idmap v sectie'] = pd.DataFrame(columns=['bestand',
                                                    'externalLocation',
                                                    'externalParameter',
                                                    'internalLocation',
                                                    'internalParameter',
                                                    ])
for idmap, idmap_subsecs in idmap_sections.items():
    for section_type, sections in idmap_subsecs.items():
        for section in sections:
            if section_type == 'KUNSTWERKEN':
                prefix = 'KW'
            if section_type == 'WATERSTANDLOCATIES':
                prefix = 'OW'
            idmap_wrong_section = [idmap for idmap in xml_to_dict(config.IdMapFiles[idmap],**section)['idMap']['map'] 
                                   if not idmap['internalLocation'][0:2] == prefix]
            if len(idmap_wrong_section):
                section_start = section['section_start'] if 'section_start' in section.keys() else ''
                section_end = section['section_end'] if 'section_end' in section.keys() else ''
                logging.warning('{} internalLocations anders dan {}XXXXXX tussen {} en {} in {}'.format(len(idmap_wrong_section),
                                                                                                       prefix,
                                                                                                       section_start,
                                                                                                       section_end,
                                                                                                       idmap))
                df = pd.DataFrame(idmap_wrong_section)
                df['bestand'] = idmap
                config_df['idmap v sectie'] = pd.concat([config_df['idmap v sectie'], df], axis=0)
    
    summary['idmaps in verkeerde sectie'] = len(config_df['idmap v sectie'])

#%% inlezen hist tags & ignore lijst
hist_tags_df = pd.read_csv(hist_tags_csv,
                           parse_dates = ['total_min_start_dt', 'total_max_end_dt'],
                           sep = ';')


#filteren hist_tags op alles wat niet in ignored staat
hist_tags_df = hist_tags_df[~hist_tags_df['serie'].isin(config_df['histTag_ignore']['UNKNOWN_SERIE'])]

#%% toevoegen lijsten met fews_locids aan his_tags_df        
hist_tags_df['fews_locid'] = hist_tags_df.apply(idmap2tags, axis=1)

#%% wegschrijven his-tags die niet zijn gematched
hist_tags_no_match_df = hist_tags_df[hist_tags_df['fews_locid'].isna()]
hist_tags_no_match_df = hist_tags_no_match_df.drop('fews_locid',axis=1)
hist_tags_no_match_df.columns = ['UNKNOWN_SERIE','STARTDATE','ENDDATE']
hist_tags_no_match_df = hist_tags_no_match_df.set_index('UNKNOWN_SERIE')
config_df['histTags_noMatch'] = hist_tags_no_match_df
summary['histTags_noMatch'] = len(hist_tags_no_match_df)

if not config_df['histTags_noMatch'].empty:
    logging.warning('{} histTags zijn niet gematched'.format(len(config_df['histTags_noMatch'])))
else:
    logging.info('alle histTags zijn gematched in idmap')



#%% aanmaken van mpt_df vanuit de fews_locid lijsten in hist_tags_df
hist_tags_df = hist_tags_df[hist_tags_df['fews_locid'].notna()]
mpt_hist_tags_df = hist_tags_df.explode('fews_locid').reset_index(drop=True)

#%% bepalen minimale start en maximale eindtijd per fews_locid. 
mpt_df = pd.concat([mpt_hist_tags_df.groupby(['fews_locid'], sort=False)['total_min_start_dt'].min(),
                    mpt_hist_tags_df.groupby(['fews_locid'], sort=False)['total_max_end_dt'].max()],
                   axis=1)

mpt_df = mpt_df.sort_index(axis=0)
mpt_df.columns = ['STARTDATE','ENDDATE']
mpt_df.index.name = 'LOC_ID'

#%% alle hoofdlocaties waar geen histag op binnekomt toevoegen
kw_locs = list(mpt_df[mpt_df.index.str.contains('KW', regex=False)].index)
h_locs = np.unique(['{}0'.format(loc[0:-1]) for loc in kw_locs])
h_locs_missing = [loc for loc in h_locs if not loc in list(mpt_df.index)]
h_locs_df = pd.DataFrame(data={'LOC_ID' : h_locs_missing,
                               'STARTDATE' : [pd.NaT]*len(h_locs_missing),
                               'ENDDATE' :  [pd.NaT]*len(h_locs_missing)})
h_locs_df = h_locs_df.set_index('LOC_ID')

mpt_df = pd.concat([mpt_df,h_locs_df],axis=0)
#%% de start en eindtijd op de hoofdlocatie updaten met de min/max van de sublocatie
def update_hlocs(row):
    loc_id = row.name
    start_date = row['STARTDATE']
    end_date = row['ENDDATE']
    
    if loc_id in h_locs:
        start_date = mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]['STARTDATE'].dropna().min()
        end_date = mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]['ENDDATE'].dropna().max()
    
    return start_date, end_date 

mpt_df[['STARTDATE','ENDDATE']] = mpt_df.apply(update_hlocs,axis=1,result_type="expand")

mpt_df = mpt_df.sort_index()
config_df['mpt'] = mpt_df

#%% consistentie parameters: zijn alle interne parameters opgenomen in parameters.xml
config_df['dubbele idmaps'] = pd.DataFrame(columns=['bestand',
                                                    'externalLocation',
                                                    'externalParameter',
                                                    'internalLocation',
                                                    'internalParameter'])
                      
for idmap_file in idmap_files:
    idmap_doubles = [id_map for id_map in idmap_dict[idmap_file] if idmap_dict[idmap_file].count(id_map) > 1]
    if len(idmap_doubles) > 0:
        idmap_doubles = list({idmap['externalLocation']:idmap for idmap in idmap_doubles}.values())
        df = pd.DataFrame(idmap_doubles,columns=['internalLocation','externalLocation','internalParameter','externalParameter'])
        df['bestand'] = idmap_file
        config_df['dubbele idmaps'] = pd.concat([config_df['dubbele idmaps'], df], axis=0)
        logging.warning('{} dubbele idmap(s) in {}'.format(len(idmap_doubles),idmap_file))
    else:
        logging.info('geen dubbele idmaps in {}'.format(idmap_file))

    summary['dubbele idmaps {}'.format(idmap_file)] = len(idmap_doubles)

#%% consistentie parameters: zijn alle interne parameters opgenomen in parameters.xml
config_parameters = list(config.get_parameters(dict_keys='parameters').keys())
id_map_parameters = [id_map['internalParameter'] for id_map in idmap_total]
params_missing = [parameter for parameter in id_map_parameters 
                  if not parameter in config_parameters]

summary['missende parameters'] = len(params_missing)

if len(params_missing) == 0:
    logging.info('alle parameters in idMaps zijn opgenomen in config')
else:
    logging.warning('{} uit idMaps missen in config'.format(len(params_missing)))
    config_df['params_missing'] =  pd.DataFrame({'parameters': params_missing})
    config_df['params_missing'] = config_df['params_missing'].set_index('parameters')

#%% wegschrijven naar excel
    
#lees input xlsx en gooi alles weg behalve de fixed_sheets
book = load_workbook(consistency_out)
for worksheet in book.worksheets:
    if not worksheet.title in fixed_sheets:
        book.remove(worksheet)

# voeg samenvatting toe
worksheet = book.create_sheet('samenvatting',1)
worksheet.sheet_properties.tabColor = '92D050'
worksheet.append(['controle','aantal'])
for cell in worksheet['{}'.format(worksheet.max_row)]:
    cell.font = Font(bold=True)
    
for key, value in summary.items():
    worksheet.append([key,value])
    if value > 0:
       worksheet[worksheet.max_row][1].fill = PatternFill(fgColor='FF0000', fill_type='solid')
    else:
        worksheet[worksheet.max_row][1].fill = PatternFill(fgColor='92D050', fill_type='solid')

worksheet.column_dimensions['A'].width=40
worksheet.auto_filter.ref = worksheet.dimensions


xls_writer = pd.ExcelWriter(consistency_out, engine='openpyxl')
xls_writer.book = book

for sheet_name, df in config_df.items():
        if (not sheet_name in fixed_sheets) & (not df.empty):
            if df.index.name == None:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=True)
            worksheet = xls_writer.sheets[sheet_name]
            for col in worksheet.columns:
                worksheet.column_dimensions[col[0].column_letter].width = 20
            worksheet.auto_filter.ref = worksheet.dimensions
            if not df.empty:
                if (sheet_name in warning_sheets):
                    worksheet.sheet_properties.tabColor = 'FF0000'
                else:
                    worksheet.sheet_properties.tabColor = '92D050'
                    
xls_writer.book.active = xls_writer.book['samenvatting']

xls_writer.save()